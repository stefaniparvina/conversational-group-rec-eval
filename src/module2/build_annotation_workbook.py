"""Module 2: build the human-annotation materials for the 15 validation transcripts
-- one readable .txt per transcript plus an Excel workbook (Guidelines + four metric
sheets with dropdowns matching the judge's output schema) for a human to fill in.
The human scores are later compared with the GPT-4o judge (validate_judge.py) to
estimate agreement. Run select_validation_set.py first."""

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# -- Configuration -------------------------------------------------------------
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

DATA_DIR  = PROJECT_ROOT / "data" / "full_dataset"
VAL_DIR   = PROJECT_ROOT / "data" / "validation"
VAL_CSV   = VAL_DIR / "validation_set.csv"
TX_DIR    = VAL_DIR / "transcripts"
OUT_XLSX  = VAL_DIR / "annotation_workbook.xlsx"

FONT = "Arial"

# -- Styles --------------------------------------------------------------------
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", start_color="305496")
TITLE_FONT  = Font(name=FONT, bold=True, size=14)
H2_FONT     = Font(name=FONT, bold=True, size=11, color="305496")
BODY_FONT   = Font(name=FONT, size=10)
FIXED_FONT  = Font(name=FONT, size=10, color="595959")   # pre-filled (do not edit)
INPUT_FILL  = PatternFill("solid", start_color="FFF2CC") # light yellow = fill me
WRAP_TOP    = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT_TOP    = Alignment(horizontal="left", vertical="top", wrap_text=True)


# -- Transcript helpers --------------------------------------------------------

def detect_shifts(data: dict) -> list:
    """Find every preference change from the turn_preferences table.
    turn_preferences[agent] is the agent's suggested option per round;
    a change between consecutive rounds is a shift."""
    shifts = []
    for agent, prefs in data.get("turn_preferences", {}).items():
        for i in range(1, len(prefs)):
            if prefs[i] != prefs[i - 1]:
                shifts.append({
                    "agent": agent,
                    "from_preference": prefs[i - 1],
                    "to_preference": prefs[i],
                    "round": i + 1,          # prefs[0] is round 1
                })
    return shifts


def initial_preference(data: dict, agent: str) -> str:
    prefs = data.get("turn_preferences", {}).get(agent, [])
    return prefs[0] if prefs else "?"


def render_transcript(data: dict) -> str:
    gid     = data.get("group_id")
    config  = data.get("group_config", "?")
    agents  = data.get("agents", [])
    n       = len(agents)
    final   = data.get("final_rec", "?")
    options = data.get("restaurant_options", [])
    tp      = data.get("turn_preferences", {})

    L = []
    bar = "=" * 70
    L.append(bar)
    L.append(f"TRANSCRIPT   group_id = {gid}   |   config = {config}")
    L.append(bar)
    L.append(f"Agents: {n}")
    L.append(f"Final group recommendation: {final}")
    if options:
        L.append("Restaurant options: " + ", ".join(options))
    L.append("")

    L.append("AGENT PROFILES  (personality/tone/behaviour are ASSIGNED -- see edge rules)")
    for ag in agents:
        name = ag.get("name", "?")
        L.append(f"  {name}")
        tone = ag.get("tone")
        if tone:
            L.append(f"     tone: {' '.join(str(tone).split())}")
        pers = ag.get("personality")
        if isinstance(pers, dict):
            L.append("     personality: "
                     + ", ".join(f"{k} {float(v):.2f}" for k, v in pers.items()))
        elif pers:
            L.append(f"     personality: {pers}")
        style = ag.get("conversation_style")
        if style:
            style_lines = [ln.strip() for ln in
                           str(style).replace("**", "").splitlines() if ln.strip()]
            if style_lines:
                L.append("     behaviour:")
                for ln in style_lines:
                    L.append(f"        {ln}")
        hist = ag.get("history", {})
        if hist:
            top = sorted(hist.items(), key=lambda x: -x[1])[:3]
            L.append("     own top ratings: "
                     + ", ".join(f"{r}({v})" for r, v in top))
        L.append("")

    L.append("PREFERENCE EVOLUTION  (the option each agent suggested, by round)")
    for agent, prefs in tp.items():
        L.append(f"  {agent:<18}: " + "  ->  ".join(prefs))
    L.append("")

    L.append("-" * 70)
    L.append("CONVERSATION")
    L.append("-" * 70)
    conv = data.get("conversation", [])
    last_round = None
    for i, msg in enumerate(conv, 1):
        rnd = ((i - 1) // n) + 1 if n else 1
        if rnd != last_round:
            L.append("")
            L.append(f"--- Round {rnd} ---")
            last_round = rnd
        text = " ".join(str(msg.get("message", "")).split())
        L.append(f"[{i}] {msg.get('agent', '?')}: {text}")
    L.append("")
    return "\n".join(L)


# -- Workbook helpers ----------------------------------------------------------

def write_header(ws, headers: list, widths: list):
    for c, (title, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def mark_input(cell):
    cell.fill = INPUT_FILL
    cell.alignment = CENTER
    cell.font = BODY_FONT


def list_dv(values: str) -> DataValidation:
    """A dropdown data validation built from a comma-separated value string."""
    dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
    dv.error = "Pick a value from the dropdown list."
    dv.errorTitle = "Invalid entry"
    return dv


# -- Sheet builders ------------------------------------------------------------

GUIDELINES = [
    ("title", "Module 2 -- Human Annotation Workbook"),
    ("body",  "You are scoring 15 group-conversation transcripts by hand. These "
              "scores are compared against the GPT-4o judge to measure agreement "
              "(Cohen's kappa). Use exactly the same rules the judge is given -- "
              "they are reproduced below."),
    ("body",  "HOW TO USE: open the transcript for a group from "
              "data/validation/transcripts/transcript_<group_id>.txt, read it, "
              "then fill the yellow cells on the four metric sheets. Categorical "
              "cells have dropdowns. Judge only on what the transcript shows."),
    ("h2",    "Sheet: ProcessQuality -- score each of 4 dimensions 0-3"),
    ("body",  "preferences_heard -- Were all agents' preferred options taken up "
              "and discussed by the other agents?"),
    ("body",  "   0 = one or more preferences never engaged with (ignored, or "
              "named with no response)."),
    ("body",  "   1 = preferences named but barely discussed; at least one agent "
              "got no substantive engagement."),
    ("body",  "   2 = every preference acknowledged and most discussed with a "
              "brief reason, but engagement uneven."),
    ("body",  "   3 = every preference explicitly discussed by others with "
              "substantive reasons; no agent left out."),
    ("body",  "shifts_justified -- When agents changed their stated preference, "
              "was the change explained?"),
    ("body",  "   0 = at least one agent changed preference with no stated reason."),
    ("body",  "   1 = shifts happened but only weakly justified (\"fine, okay\", "
              "bare capitulation, no content)."),
    ("body",  "   2 = most shifts had a clear content-based reason; a minority "
              "unexplained or purely social."),
    ("body",  "   3 = every shift had an explicit, content-based reason (a "
              "concrete merit/drawback of an option)."),
    ("body",  "   If NO agent shifted: do not score 0 or 3 by default -- score 2 "
              "if the group genuinely weighed options and converged on reasoned "
              "grounds, lower if it converged with little discussion."),
    ("body",  "mutual_respect -- Was the tone respectful toward agents and their "
              "preferences throughout?"),
    ("body",  "   0 = personal attacks, mocking, or repeated dismissiveness."),
    ("body",  "   1 = generally civil but with clear dismissiveness (a preference "
              "brushed off, a curt rejection)."),
    ("body",  "   2 = respectful throughout; disagreement expressed politely; "
              "minor curtness at most."),
    ("body",  "   3 = consistently respectful; disagreement paired with "
              "acknowledgement of the other view."),
    ("body",  "logical_support -- Were positions backed by specific arguments "
              "about the restaurants?"),
    ("body",  "   0 = choices asserted with no reasons (just \"I want X\")."),
    ("body",  "   1 = reasons given but vague/generic (\"it's nice\"), no "
              "restaurant-specific content."),
    ("body",  "   2 = most claims backed by a concrete feature (rating, cuisine, "
              "price, location)."),
    ("body",  "   3 = claims consistently backed by specific, comparative "
              "arguments weighing more than one option."),
    ("body",  "If a transcript falls between two anchors, choose the LOWER one."),
    ("h2",    "Sheet: MentionRate -- one row per agent"),
    ("body",  "acknowledged -- Yes if that agent's preferred option received at "
              "least one non-dismissive mention from another agent; otherwise No."),
    ("body",  "dominant_sentiment -- the overall tone of how others treated that "
              "agent's preference: positive (supported / seriously considered), "
              "neutral (mentioned without evaluation), dismissive (rejected, "
              "mocked, argued down). If unclear, choose neutral. If the "
              "preference was never mentioned at all, leave sentiment blank."),
    ("h2",    "Sheet: JustifiedShifts -- one row per detected preference shift"),
    ("body",  "Each shift (agent, from, to, round) is pre-filled. Classify the "
              "shift_type: quality-based (agent cited specific merits/drawbacks), "
              "social (group pressure, solidarity, conflict avoidance), "
              "unexplained (no reason given). If it could be quality-based or "
              "social and the transcript does not make it clear, choose "
              "unexplained."),
    ("h2",    "Sheet: RepetitionIndex -- one row per agent"),
    ("body",  "times_repeated_before_ack -- how many times the agent restated "
              "their initial preference before another agent substantively "
              "acknowledged it. 0 = acknowledged on first mention. If never "
              "acknowledged, count the total number of restatements."),
    ("h2",    "Edge cases (apply consistently)"),
    ("body",  "- Assigned personality is not disrespect: a blunt or assertive "
              "assigned style is not by itself a respect violation -- judge "
              "respect by how agents treat each other's preferences."),
    ("body",  "- Short conversations: score only what is present; do not infer "
              "unobserved respect, reasoning, or engagement. Absence of evidence "
              "scores at the lower anchor, never assumed positive."),
    ("body",  "- Use the 'notes' column whenever a case felt borderline -- this "
              "helps interpret any judge-vs-human disagreements later."),
]


def build_guidelines(ws):
    ws.sheet_properties.tabColor = "305496"
    ws.column_dimensions["A"].width = 118
    r = 1
    for style, text in GUIDELINES:
        cell = ws.cell(row=r, column=1, value=text)
        cell.alignment = LEFT_TOP
        if style == "title":
            cell.font = TITLE_FONT
            ws.row_dimensions[r].height = 22
        elif style == "h2":
            cell.font = H2_FONT
            ws.row_dimensions[r].height = 20
        else:
            cell.font = BODY_FONT
            ws.row_dimensions[r].height = 30
        r += 1


def build_process_quality(ws, transcripts):
    headers = ["group_id", "group_config",
               "preferences_heard", "shifts_justified",
               "mutual_respect", "logical_support", "notes"]
    widths  = [11, 14, 17, 17, 16, 16, 48]
    write_header(ws, headers, widths)

    dv = list_dv("0,1,2,3")
    ws.add_data_validation(dv)

    for i, (gid, data) in enumerate(transcripts):
        row = i + 2
        ws.cell(row=row, column=1, value=gid).font = FIXED_FONT
        ws.cell(row=row, column=2,
                value=data.get("group_config", "?")).font = FIXED_FONT
        for col in (3, 4, 5, 6):                 # the four 0-3 score columns
            mark_input(ws.cell(row=row, column=col))
            dv.add(ws.cell(row=row, column=col))
        ws.cell(row=row, column=7).font = BODY_FONT
        ws.cell(row=row, column=7).alignment = WRAP_TOP


def build_mention_rate(ws, transcripts):
    headers = ["group_id", "agent", "initial_preference",
               "acknowledged", "dominant_sentiment", "notes"]
    widths  = [11, 18, 17, 15, 19, 46]
    write_header(ws, headers, widths)

    dv_ack  = list_dv("Yes,No")
    dv_sent = list_dv("positive,neutral,dismissive")
    ws.add_data_validation(dv_ack)
    ws.add_data_validation(dv_sent)

    row = 2
    for gid, data in transcripts:
        for ag in data.get("agents", []):
            name = ag.get("name", "?")
            ws.cell(row=row, column=1, value=gid).font = FIXED_FONT
            ws.cell(row=row, column=2, value=name).font = FIXED_FONT
            ws.cell(row=row, column=3,
                    value=initial_preference(data, name)).font = FIXED_FONT
            mark_input(ws.cell(row=row, column=4)); dv_ack.add(ws.cell(row=row, column=4))
            mark_input(ws.cell(row=row, column=5)); dv_sent.add(ws.cell(row=row, column=5))
            ws.cell(row=row, column=6).font = BODY_FONT
            ws.cell(row=row, column=6).alignment = WRAP_TOP
            row += 1


def build_justified_shifts(ws, transcripts):
    headers = ["group_id", "agent", "from_preference",
               "to_preference", "round", "shift_type", "notes"]
    widths  = [11, 18, 16, 16, 9, 17, 44]
    write_header(ws, headers, widths)

    dv = list_dv("quality-based,social,unexplained")
    ws.add_data_validation(dv)

    row = 2
    n_shifts = 0
    for gid, data in transcripts:
        for s in detect_shifts(data):
            ws.cell(row=row, column=1, value=gid).font = FIXED_FONT
            ws.cell(row=row, column=2, value=s["agent"]).font = FIXED_FONT
            ws.cell(row=row, column=3, value=s["from_preference"]).font = FIXED_FONT
            ws.cell(row=row, column=4, value=s["to_preference"]).font = FIXED_FONT
            ws.cell(row=row, column=5, value=s["round"]).font = FIXED_FONT
            mark_input(ws.cell(row=row, column=6)); dv.add(ws.cell(row=row, column=6))
            ws.cell(row=row, column=7).font = BODY_FONT
            ws.cell(row=row, column=7).alignment = WRAP_TOP
            row += 1
            n_shifts += 1
    if n_shifts == 0:
        ws.cell(row=2, column=1,
                value="(no preference shifts detected in any transcript)"
                ).font = FIXED_FONT
    return n_shifts


def build_repetition_index(ws, transcripts):
    headers = ["group_id", "agent", "initial_preference",
               "times_repeated_before_ack", "notes"]
    widths  = [11, 18, 17, 26, 50]
    write_header(ws, headers, widths)

    dv = DataValidation(type="whole", operator="greaterThanOrEqual",
                        formula1="0", allow_blank=True)
    dv.error = "Enter a whole number of 0 or greater."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)

    row = 2
    for gid, data in transcripts:
        for ag in data.get("agents", []):
            name = ag.get("name", "?")
            ws.cell(row=row, column=1, value=gid).font = FIXED_FONT
            ws.cell(row=row, column=2, value=name).font = FIXED_FONT
            ws.cell(row=row, column=3,
                    value=initial_preference(data, name)).font = FIXED_FONT
            mark_input(ws.cell(row=row, column=4)); dv.add(ws.cell(row=row, column=4))
            ws.cell(row=row, column=5).font = BODY_FONT
            ws.cell(row=row, column=5).alignment = WRAP_TOP
            row += 1


# -- Main ----------------------------------------------------------------------

def main():
    if not VAL_CSV.exists():
        raise FileNotFoundError(
            f"{VAL_CSV} not found -- run select_validation_set.py first.")

    val = pd.read_csv(VAL_CSV)
    gids = val["group_id"].tolist()
    print(f"Validation set: {len(gids)} transcripts -> {gids}")

    transcripts = []
    for gid in gids:
        path = DATA_DIR / f"group_simulation_{gid}.json"
        if not path.exists():
            raise FileNotFoundError(f"transcript file missing: {path}")
        transcripts.append((gid, json.loads(path.read_text(encoding="utf-8"))))

    TX_DIR.mkdir(parents=True, exist_ok=True)
    for gid, data in transcripts:
        (TX_DIR / f"transcript_{gid}.txt").write_text(
            render_transcript(data), encoding="utf-8")
    print(f"Wrote {len(transcripts)} transcript files -> {TX_DIR}")

    wb = Workbook()
    build_guidelines(wb.active)
    wb.active.title = "Guidelines"

    build_process_quality(wb.create_sheet("ProcessQuality"), transcripts)
    build_mention_rate(wb.create_sheet("MentionRate"), transcripts)
    n_shifts = build_justified_shifts(wb.create_sheet("JustifiedShifts"), transcripts)
    build_repetition_index(wb.create_sheet("RepetitionIndex"), transcripts)

    VAL_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    n_agents = sum(len(d.get("agents", [])) for _, d in transcripts)
    print(f"Saved annotation workbook -> {OUT_XLSX}")
    print(f"  ProcessQuality : {len(transcripts)} rows")
    print(f"  MentionRate    : {n_agents} rows")
    print(f"  JustifiedShifts: {n_shifts} rows")
    print(f"  RepetitionIndex: {n_agents} rows")


if __name__ == "__main__":
    main()
